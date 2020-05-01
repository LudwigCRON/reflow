#!/usr/bin/env python3
# coding: utf-8

import os
import sys
import argparse

from openpyxl import load_workbook
from mako.template import Template

sys.path.append(os.environ["REFLOW"])

import common.utils as utils


def parse_time(s: str):
    """
    convert time format '<value> <unit>' in a floating point value
    (see utils.parse_eng_unit)

    Args:
        s (str): string to be convert
    Returns:
        1e-12 (default) if s is not string
        the value scaled by the unit
    """
    return utils.parse_eng_unit(s, base_unit='s', default=1e-12)


def parse_cap(s):
    """
    convert capacitiance format '<value> <unit>' in a floating point value
    (see utils.parse_eng_unit)

    Args:
        s (str): string to be convert
    Returns:
        1e-15 (default) if s is not string
        the value scaled by the unit
    """
    return utils.parse_eng_unit(s, base_unit='f', default=1e-15)


def parse_name(s):
    if not isinstance(s, str):
        return None
    start = s.find("[") if "[" in s else s.find("<")
    end = s.find("]") if "]" in s else s.find(">")
    sep = s.find(":")
    # just bracket return it
    if start < 1:
        return s, 1
    # if bus
    if sep > 0:
        bounds = s[start + 1:end].split(":")
        bounds = [int(b, 10) for b in bounds]
        return s[:start], abs(bounds[0] - bounds[-1]) + 1
    else:
        return s[:start], 1


class Pins(object):
    __slots__ = [
        "ana_name",
        "dig_name",
        "width",
        "type",
        "direction",
        "capacitance",
        "transition",
        "access",
        "delay",
        "setup",
        "hold"
    ]

    def __init__(self):
        self.access = {}
        self.delay = {}
        self.setup = {}
        self.hold = {}

    def __repr__(self):
        s = ["%s: %s" % (s, getattr(self, s, '')) for s in Pins.__slots__]
        return '\n'.join(s)

    def to_dict(self):
        return {k: getattr(self, k) for k in self.__slots__}

    @staticmethod
    def parse_dict(d):
        p = Pins()
        # get most probable key name
        for k in d.keys():
            # filter
            if k is None:
                continue
            # parse
            key = k.lower()
            if "ana" in key and "name" in key:
                p.ana_name, p.width = parse_name(d.get(k, "%s_%d"))
            elif "name" in key:
                p.dig_name, p.width = parse_name(d.get(k, "ms_%s_%d"))
            elif "dir" in key:
                t = d.get(k, "dig->ana").strip().lower()
                p.direction = "input" if t.startswith("dig") else "output"
            elif "type" in key:
                p.type = d.get(k, "UNCONSTRAINED")
            elif "access" in key:
                corner = key.replace("access", '').strip()
                dt = parse_time(d.get(k, "1ns"))
                p.access[corner if corner else 'typ'] = dt
            elif "setup" in key:
                corner = key.replace("setup", '').strip()
                dt = parse_time(d.get(k, "1ns"))
                p.setup[corner if corner else 'typ'] = dt
            elif "hold" in key:
                corner = key.replace("hold", '').strip()
                dt = parse_time(d.get(k, "1ns"))
                p.hold[corner if corner else 'typ'] = dt
            elif "delay" in key:
                corner = key.replace("delay", '').strip()
                dt = parse_time(d.get(k, "1ns"))
                p.delay[corner if corner else 'typ'] = dt
            elif "trans" in key:
                trans = parse_time(d.get(k, "1ns"))
                if trans:
                    p.transition = trans
            elif "cap" in key:
                cap = parse_cap(d.get(k, "10fF"))
                if cap:
                    p.capacitance = cap
        # retrive value or set default
        return p


def read_timings_row(row, headers=[]):
    values = [cell.value for cell in row]
    if all([v is None for v in values]):
        return None
    d = dict(zip(headers, values))
    return Pins.parse_dict(d)


def read_timings(ws):
    pins = []
    header = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            header = [h.value for h in row]
            continue
        pin = read_timings_row(row, header)
        if pin:
            pins.append(pin)
    return pins


def read_description_row(row, headers=[]):
    values = [cell.value for cell in row]
    if all([v is None for v in values]):
        return None
    return dict(zip(headers, values))


def read_description(ws):
    description, headers = {}, []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            headers = [h.value for h in row]
            continue
        elif i == 1:
            description = read_description_row(row, headers)
        else:
            d = read_description_row(row, headers)
            for k, v in d.items():
                if isinstance(description[k], list):
                    description[k].append(v)
                elif v is not None:
                    description[k] = [description[k], v]
    # list corners
    corners = [k for k in description.keys() if "corner" in k.lower()]
    # reshape corners
    a = {}
    for corner in corners:
        corn = corner.lower().replace("corner", '').strip()
        a[corn] = dict(zip(["process", "voltage", "temperature"], description.pop(corner)))
    description["corners"] = a
    # normalize
    _ = [('_'.join(k.lower().split(' ')), k) for k in description.keys()]
    for new_k, k in _:
        description[new_k] = description.pop(k)
    return description


class Lib:
    __slots__ = ["name", "area", "pins", "types", "corners"]

    def __init__(self, desc, pins):
        self.name = desc.get("name_of_the_cell")
        self.area = "block_area_(um2)"
        self.pins = pins
        self.types = [pin for pin in pins if pin.width > 1]
        self.corners = []
        for corner, condition in desc.get("corners", {}).items():
            self.corners.append({
                "name": corner,
                "condition": condition,
                "library": "%s_%s_%sV_%sC" % (
                    desc.get("name_of_the_cell"),
                    corner,
                    ("%.2f" % condition.get("voltage")).replace('.', '_'),
                    str(condition.get("temperature")).replace('-', 'm')
                )
            })

    def to_dict(self):
        return {k: getattr(self, k) for k in self.__slots__}


def create_libs(desc: dict, pins: list, output_dir: str, verbose: bool = False):
    """
    generate a lib file for each corners
    """
    db = desc.copy()
    db["block_name"] = desc["name_of_the_cell"]
    db["area"] = db.pop("block_area_(um2)")
    db["pins"] = pins
    db["types"] = [pin for pin in pins if pin.width > 1]
    lib_paths = []
    for corner, condition in desc.get("corners", {}).items():
        db["library"] = "%s_%s_%sV_%sC" % (
            desc.get("name_of_the_cell"),
            corner,
            ("%.2f" % condition.get("voltage")).replace('.', '_'),
            str(condition.get("temperature")).replace('-', 'm')
        )
        db["corner_name"] = corner
        db["corner"] = condition
        if verbose:
            print(db)
        # create directory if does not exist
        os.makedirs(output_dir, exist_ok=True)
        # generate lib file
        template_file = os.path.join(os.path.dirname(__file__), "./template_ana.lib.mako")
        _tmp = Template(filename=template_file)
        lib_path = os.path.join(output_dir, "%s.lib" % db["library"])
        with open(lib_path, "w+") as fp:
            fp.write(_tmp.render_unicode(**db))
        lib_paths.append(lib_path)
    return lib_paths


def main(file, output_dir, verbose: bool = False):
    wb = load_workbook(file, data_only=True)
    # read corners and extraction informations
    ws = wb["General"]
    desc = read_description(ws)
    # read pins and their timings
    ws = wb["Timing"]
    pins = read_timings(ws)
    # create a lib for each corners
    lib = Lib(desc, pins)
    libs = create_libs(desc, pins, output_dir, verbose)
    return libs, lib


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="convert excel file into lib")
    parser.add_argument("-i", "--input", help="excel input file")
    parser.add_argument("-o", "--output", help="output lib file")
    parser.add_argument("-v", "--verbose", action="store_true", default=False)
    args = parser.parse_args()

    main(args.input, args.output, args.verbose)
